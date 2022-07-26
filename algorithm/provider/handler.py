import math
from provider.model.greedy_pav import *
import json
import csv
from itertools import combinations, chain
import os
import openpyxl
from pathlib import Path
from typing import List


def powerset(iterable):
    s = list(iterable)
    return chain.from_iterable(combinations(s, r) for r in range(len(s) + 1))


def voter_agree(voters: List[Voter]) -> list:
    res = []
    count = 0
    v = voters[0]
    for c in v.preferences:
        for voter_preferences in voters:
            if c in voter_preferences.preferences:
                count += 1
        if count == len(voters):
            res.append(c)
        count = 0
    return res


def Global_Justified_Representation(X_result: dict, voters: List[Voter], offices_candidates: dict):
    X = [Candidate(j, i) for i, j in X_result.items()]
    s = f'''<h1>Global Justified Representation:</h1>
<h2>Definition:</h2>
<p>V:set of all voters. <br />n:number of voters.<br />k:number of offices.<br />
  X:office allocation.<br />Aj:all the candidates that run to office j.<br />
  A:∪Aj all the candidates.<br />X is said to satisfy Global Justified Representation if:<br />
  For each subgroup of voters V'⊆ V that its size >= n/k and for all Aj:<br />
  Define c_V': set of all the candidates that all the voters in V' agree on.<br />
  if (c_V' ∩ Aj) !=∅ then (X ∩ c_V') !=∅.<br /><br />
  <h2>Global Proportional Justified Representation</h2>
  if (c_V' ∩ Aj) !=∅  then |(X ∩ c_V')| ≥⌊(|V'|)/(|V|)k⌋ <br /><br />
  GreedyPAV satisfies GJR but not always GPJR (there exists an algorithm that satisfies GPJR called FPT, but it is running in super-polynomial time )
  <h2>Prof for this result:</h2>
  <h3>All the subgroup that we discuss for this results:</h3></p>
    '''
    s += '''<table width="300" border="1" cellpadding="5" style="text-align: center">
            <thead>
            <tr>
            <td>Office</td>
            <td>Candidate</td>
            </tr>
            </thead>
            <tbody>'''
    for o, c in X_result.items():
        s += f"""<tr><td>{o}</td><td>{c}</td></tr>"""
    s += "</tbody></table>"
    s += f'''<p>Note:the subgroup that display are only those that big enough
     only the subgroup that  its size >= n/k={len(voters) / len(offices_candidates.keys())}</p>
    <table width="300" border="1" cellpadding="5" style="text-align: center">
            <thead>
            <tr>
            <td>Voters subgroup</td>
            <td>Candidate they agree on</td>
            <td>Candidate who won</td>
            <td>satisfies GPJR</td>
            </tr>
            </thead>
            <tbody>'''
    V = list(powerset(voters))
    for v in V:
        if len(v) >= (len(voters) / len(offices_candidates.keys())):
            v_agree = voter_agree(v)
            for office, A_j in offices_candidates.items():
                if len(set(v_agree).intersection(A_j)) > 0:
                    t = list(set(v_agree).intersection(X))
                    if len(t) > 0:
                        lt = [i.name for i in t]
                        vs = [i.name for i in v]
                        cs = ' '.join(lt)
                        if len(t) >= math.floor(len(v) / len(voters) * len(offices_candidates.keys())):
                            s += f"""<tr><td>{' '.join(vs)}</td><td>{list(set(v_agree).intersection(A_j))[0].name}</td><td>{cs}</td><td>&#10004;</td></tr>"""
                        else:
                            s += f"""<tr><td>{' '.join(vs)}</td><td>{list(set(v_agree).intersection(A_j))[0].name}</td><td>{cs}</td><td>&#10006;</td></tr>"""
                        print(f"{t} were elected which they also agree on.")
                    else:
                        print("bad")
        else:
            pass
    s += "</tbody></table>"
    return s


def from_json(json_res: str):
    res_dict = json.loads(json_res)
    candidates = res_dict['candidates']
    voters = res_dict['voters']
    offices_candidates = {}
    for candidate_list in candidates.values():
        for candidate in candidate_list:
            c = Candidate(candidate['name'], candidate['office'])
            if c.office not in offices_candidates:
                offices_candidates[c.office] = []
            offices_candidates[c.office].append(c)
    voter_list = []
    for voter in voters:
        preferences = []
        for pref in voter["preferences"]:
            preferences.append(Candidate(**pref))
        v = Voter(preferences)
        print(v)
        voter_list.append(v)

    return offices_candidates, voter_list


def convert_request(offices: list = [], candidates: list = [], voters: list = [], votersNames: list = []):
    offices_candidates = {}
    voter_list = []
    num_to_office = {i: offices[i] for i in range(len(offices))}
    for office_num in range(len(candidates)):
        for candidate in candidates[office_num]:
            c = Candidate(candidate, num_to_office[office_num])
            if c.office not in offices_candidates:
                offices_candidates[c.office] = []
            offices_candidates[c.office].append(c)
    for voter in range(len(voters)):
        preferences = []
        for i in range(len(voters[voter])):
            preferences.append(Candidate(voters[voter][i], num_to_office[i]))
        v = Voter(preferences=preferences, name=votersNames[voter])
        voter_list.append(v)
    return offices_candidates, voter_list


def from_csv(file):
    row = []
    with open(file) as f:
        csv_reader = csv.reader(f)
        h = next(csv_reader)
    print(h)


def define_result(dic_res: dict = {}) -> list:
    res = []
    for office, winner in dic_res.items():
        res.append(winner)
    return res


def voter_to_winner(dic_res: dict = {}, voter_list: list = []) -> list:
    res = {}
    for winner in dic_res.values():
        res[winner] = []
        for v in voter_list:
            for c in v.preferences:
                if c.name == winner:
                    res[winner].append(v.name)
    return list(res.values())


def from_xslx(file):
    offices_candidates = {}
    voter_list = []
    print(file)
    xlsx_file = Path(file)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.get_sheet_by_name('Offices & Candidates')
    print(sheet)
    col_names = []
    for column in sheet.iter_cols(min_col=3, max_col=sheet.max_column):
        for i in range(5, sheet.max_row):
            if column[i].value:
                col_names.append(column[i].value)
                c = Candidate(name=column[i].value, office=column[4].value)
                if c.office not in offices_candidates:
                    offices_candidates[c.office] = []
                offices_candidates[c.office].append(c)
    sheet = wb_obj.get_sheet_by_name('Votes')
    num_to_office = []
    for column in sheet.iter_cols(min_row=5, min_col=4, max_col=sheet.max_column):
        num_to_office.append(column[0].value)
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row):
        count = 0
        preferences = []
        flag = False
        for cell in row:
            if cell.value:
                flag = True

                preferences.append(Candidate(cell.value, num_to_office[count]))
            count += 1

        if flag:
            v = Voter(preferences)
            voter_list.append(v)

    if os.path.exists(file):
        os.remove(file)
    return offices_candidates, voter_list


def start_algo(json_res: str = None):
    print(json_res)
    if json_res:
        try:
            offices_candidates, voter_list = convert_request(json_res['offices'], json_res['candidates'],
                                                             json_res['voters'], json_res['votersNames'])
        except Exception as e:
            print(e)
    else:
        offices_candidates, voter_list = demo()
    a = greedy_PAV(voters=voter_list, offices_candidates=offices_candidates)
    winner_votes = voter_to_winner(a, voter_list)
    res = define_result(a)
    s = Global_Justified_Representation(a, voter_list, offices_candidates)
    with open("./files/explanation.html", 'w', encoding="utf-8") as f:
        f.write(s)
    return res, winner_votes


def save_result_to_csv(res: dict = {}):
    head = ['Office', 'Candidate']
    rows = [[o, c] for o, c in res.items()]
    with open("./files/result.csv", 'w') as f:
        csv_writer = csv.writer(f)
        csv_writer.writerow(head)
        csv_writer.writerows(rows)


""""
(1) Check if all offices have a name
(2) Check if there are different names for all the offices
(3) Check if there is at least one candidate for each office
(4) Check if all candidates have different names
"""


def validation_file(offices_candidates: dict = None, voters: List[Voter] = None) -> str:
    """" (1) Check if all offices have a name """
    for office_name in offices_candidates.keys():
        if not office_name:
            return "Please fill in all the names of the offices"

    """" (2) Check if there are different names for all the offices """
    office_name = set()
    for office_key in offices_candidates.keys():
        office_name.add(office_key)
    if len(offices_candidates) != len(office_name):
        return "Please enter different names for each office"

    """" (3) Check if there is at least one candidate for each office """
    for office_value in offices_candidates.values():
        if len(office_value) == 0:
            return "Please insert at least one candidate for each office"

    """" (4) Check if all candidates have different names """
    all_candidates = list()
    all_candidates_set = set()
    for office_value in offices_candidates.values():
        for candidate in office_value:
            all_candidates.append(candidate.name)
            all_candidates_set.add(candidate.name)
    if len(all_candidates) != len(all_candidates_set):
        return "Please enter different names for each candidate"

    return "success"


def start_algo_uploud(file):
    offices_candidates, voter_list = from_xslx(file)
    print(offices_candidates)
    massege = validation_file(offices_candidates, voter_list)
    if massege != 'success':
        raise Exception(massege)
    a = greedy_PAV(voters=voter_list, offices_candidates=offices_candidates)
    save_result_to_csv(a)
    return a


def demo():
    a = Candidate('a', '1')
    b = Candidate('b', '1')
    c = Candidate('c', '2')
    d = Candidate('d', '2')
    e = Candidate('e', '3')
    f = Candidate('f', '3')
    g = Candidate('g', '3')
    dict = {'1': [a, b], '2': [c, d], '3': [e, f, g]}
    v_1 = Voter([a, c, e], "Moshe")
    v_2 = Voter([a, c, f], "Dani")
    v_3 = Voter([a, d, f], "Roee")
    voters = [v_1, v_2, v_3]
    candidates = dict
    with open("../j.json", "w") as f:
        json.dump({'candidates': candidates, 'voters': voters}, fp=f, default=lambda o: o.__dict__, indent=4)
    return dict, voters


def demo2():
    a = Candidate('a', '1')
    b = Candidate('b', '1')
    c = Candidate('c', '2')
    d = Candidate('d', '2')
    e = Candidate('e', '3')
    f = Candidate('f', '3')
    g = Candidate('g', '3')
    dict = {'1': [a, b], '2': [c, d], '3': [e, f, g]}
    v_1 = Voter([a, c, e])
    v_2 = Voter([a, c, f])
    v_3 = Voter([a, d, f])
    voters = [v_1, v_2, v_3]
    candidates = dict

    s = json.dumps({'candidates': candidates, 'voters': voters}, default=lambda o: o.__dict__, indent=4)
    return s


if __name__ == '__main__':
    a = start_algo()
    print(a)

